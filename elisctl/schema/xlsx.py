import json
from contextlib import suppress
from decimal import Decimal

import click
import re
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import save_virtual_workbook
from typing import List, Optional, Dict, Callable, Union, Tuple, IO, Any

from elisctl.schema import transform

SCHEMA_SHEET_NAME = "Schema"
OPTIONS_SHEET_PREFIX = "Options of "

ATTRS_INDEX = 4

DEFAULT_ATTRIBUTE_TYPES = {
    "label": "string",
    "type": "string",
    "rir_field_names": "json",
    "format": "string",
    "default_value": "json",
    "constraints": "json",
    "hidden": "bool",
    "can_export": "bool",
    "score_threshold": "number",
    "width_chars": "number",
    "icon": "string",
    "min_occurrences": "number",
    "max_occurrences": "number",
    "use_rir_content": "bool",
}

HEADER_RE = re.compile(r"(.*) \((.*)\)")

CATEGORIES = ["section", "multivalue", "tuple", "datapoint"]

SchemaContent = List[dict]
SchemaDatapoint = Dict[str, Union[str, int, float, List, Dict]]


class SchemaToXlsx:
    def convert(self, schema_content: SchemaContent) -> bytes:
        workbook = Workbook()
        self._create_main_sheet(workbook, schema_content)
        self._create_option_sheets(workbook, schema_content)
        return save_virtual_workbook(workbook)

    def _create_main_sheet(self, workbook: Workbook, schema_content: SchemaContent) -> None:
        attribute_types = self._create_attribute_types(schema_content)

        schema_sheet = workbook.active
        schema_sheet.title = SCHEMA_SHEET_NAME
        headers = ["Section", "Multivalue", "Tuple", "Data"] + [
            f"{name} ({type_})" for name, type_ in attribute_types.items()
        ]
        schema_sheet.append(headers)
        schema_sheet.freeze_panes = "E2"

        attr_rows = _traverse_schema_in_order(
            schema_content, self._create_datapoint_row, types=attribute_types
        )
        for row in attr_rows:
            schema_sheet.append(row)
        _adjust_column_width(schema_sheet)

    def _create_datapoint_row(
        self, datapoint: Dict, parent_containers: List[dict], types: Dict
    ) -> List[Optional[str]]:
        structural_row = self._create_structural_row(datapoint, parent_containers)
        attr_values = self._datapoint_to_excel_row(datapoint, types)
        return structural_row + attr_values

    @staticmethod
    def _create_option_sheets(workbook: Workbook, schema_content: SchemaContent) -> None:
        def _options_to_excel(datapoint: dict, _) -> None:
            if "options" not in datapoint:
                return
            options_sheet = workbook.create_sheet(OPTIONS_SHEET_PREFIX + datapoint["id"])
            options_sheet.append(["Value", "Label"])
            options_sheet.freeze_panes = "A2"
            for item in datapoint["options"]:
                options_sheet.append([item["value"], item["label"]])
            _adjust_column_width(options_sheet)

        _traverse_schema_in_order(schema_content, _options_to_excel)

    def _create_attribute_types(self, schema_content: SchemaContent) -> Dict[str, str]:
        additional_attributes = _uniq(
            _traverse_schema_in_order(schema_content, self._extract_unknown_attributes)
        )
        if additional_attributes:
            click.echo(f"Warning: unknown attributes: {', '.join(additional_attributes)}", err=True)
        return {**DEFAULT_ATTRIBUTE_TYPES, **additional_attributes}

    @staticmethod
    def _extract_unknown_attributes(datapoint: Dict, _) -> Dict:
        result = {}
        for attr_name in datapoint:
            if attr_name in ("id", "category", "options", "children"):
                continue
            if attr_name not in DEFAULT_ATTRIBUTE_TYPES:
                result[attr_name] = "json"
        return result

    @staticmethod
    def _create_structural_row(
        datapoint: Dict, parent_containers: List[dict]
    ) -> List[Optional[str]]:
        structural_datapoints = parent_containers + [
            {"category": datapoint["category"], "id": datapoint["id"]}
        ]
        row = []
        for category in ("section", "multivalue", "tuple", "datapoint"):
            try:
                id_ = list(filter(lambda i: i["category"] == category, structural_datapoints))[0][
                    "id"
                ]
            except IndexError:
                id_ = None
            row.append(id_)
        return row

    @staticmethod
    def _datapoint_to_excel_row(datapoint: dict, attribute_types: dict) -> List[Optional[str]]:
        def _value_to_excel(key: str):
            try:
                value = datapoint[key]
            except KeyError:
                return None
            if attribute_types[key] in ("json", "bool"):
                return json.dumps(value)
            return value

        return [_value_to_excel(key) for key in attribute_types]


class XlsxToSchema:
    def convert(self, xlsx: IO[bytes]) -> SchemaContent:
        workbook = load_workbook(filename=xlsx)
        schema = self._construct_objects_from_workbook(workbook)
        return self._enhance_datapoints_with_options(workbook, schema)

    def _construct_objects_from_workbook(self, workbook: Workbook) -> SchemaContent:
        schema_sheet = workbook[SCHEMA_SHEET_NAME]
        sheet_rows = ([_safe_strip(cell) for cell in columns] for columns in schema_sheet.values)
        attribute_types = self._extract_attribute_types(next(sheet_rows)[ATTRS_INDEX:])

        schema: SchemaContent = []
        for columns in sheet_rows:
            id_, parent_id, category, attrs = self._extract_attributes(columns)

            schema = transform.traverse_datapoints(
                schema,
                transform.add,
                parent_id=parent_id,
                datapoint_to_add=self._excel_row_to_datapoint(
                    id_, category, attrs, attribute_types
                ),
            )

        return schema

    @staticmethod
    def _extract_attributes(
        columns: List[Optional[str]]
    ) -> Tuple[str, Optional[str], str, List[Optional[str]]]:
        ids, attrs = columns[:ATTRS_INDEX], columns[ATTRS_INDEX:]

        def get_id() -> Tuple[str, Optional[str]]:
            valid_ids: List[str] = [id__ for id__ in ids if id__ is not None]
            id__ = valid_ids[-1]
            try:
                parent_id_: Optional[str] = valid_ids[-2]
            except IndexError:
                parent_id_ = None
            return id__, parent_id_

        def get_category(id__: str) -> str:
            return CATEGORIES[ids.index(id__)]

        id_, parent_id = get_id()
        return id_, parent_id, get_category(id_), attrs

    @staticmethod
    def _enhance_datapoints_with_options(
        workbook: Workbook, schema: SchemaContent
    ) -> SchemaContent:
        for sheet_name in workbook.sheetnames:
            if not sheet_name.startswith(OPTIONS_SHEET_PREFIX):
                continue
            id_ = sheet_name[len(OPTIONS_SHEET_PREFIX) :]
            sheet = workbook[sheet_name]
            sheet_rows = sheet.values
            next(sheet_rows)
            options = [{"value": item[0], "label": item[1]} for item in sheet_rows]
            schema = transform.traverse_datapoints(
                schema, transform.substitute_options, id_=id_, options=options
            )
        return schema

    @staticmethod
    def _extract_attribute_types(attr_headers: List[str]) -> Dict[str, str]:
        result = {}
        for header in attr_headers:
            match = HEADER_RE.match(header)
            if not match:
                raise click.ClickException(f"Invalid header format: {header}")
            result[match.group(1)] = match.group(2)
        return result

    def _excel_row_to_datapoint(
        self, id_: str, category: str, attrs: List[Optional[str]], attribute_types: Dict[str, str]
    ) -> SchemaDatapoint:
        all_attrs = (
            (name, self._deserialize_value(value, type_))
            for value, (name, type_) in zip(attrs, attribute_types.items())
        )
        valid_attrs = {name: value for name, value in all_attrs if value is not None}
        datapoint: SchemaDatapoint = {"id": id_, "category": category, **valid_attrs}
        if datapoint["category"] in ("section", "tuple"):
            datapoint["children"] = []
        elif datapoint["category"] == "multivalue":
            datapoint["children"] = {}
        return datapoint

    @staticmethod
    def _deserialize_value(value: Optional[str], type_: str):
        if value == "" or value is None:
            return None
        if type_ == "string":
            return str(value)
        if type_ == "number":
            try:
                return int(value)
            except ValueError:
                pass
            try:
                return Decimal(value)
            except ValueError:
                raise click.ClickException(f"Cannot convert to number: {value}")
        if type_ == "bool":
            try:
                return json.loads(value)
            except TypeError:
                raise click.ClickException(f"Cannot convert to bool: {value}")
        if type_ == "json":
            try:
                return json.loads(value)
            except TypeError:
                raise click.ClickException(f"Cannot convert to json: {value}")
        click.echo(f"Unknown value type: {type_}", err=True)


def _traverse_schema_in_order(
    datapoints: List[dict], callback: Callable, parent_containers: List[Dict] = None, **kwargs
) -> List:
    parent_containers = parent_containers or []
    result = []
    for datapoint in datapoints:
        result.append(callback(datapoint, parent_containers, **kwargs))
        if datapoint["category"] == "datapoint":
            continue
        children = datapoint.get("children", [])
        if not isinstance(children, list):
            children = [children]
        result.extend(
            _traverse_schema_in_order(
                children,
                callback,
                parent_containers + [{"category": datapoint["category"], "id": datapoint["id"]}],
                **kwargs,
            )
        )
    return result


def _uniq(l: List[Dict]) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for item in l:
        result.update(**item)
    return result


def _safe_strip(value: Any) -> Any:
    with suppress(AttributeError):
        return value.strip()
    return value


# Implemented according to https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
def _adjust_column_width(worksheet: Worksheet) -> None:
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except KeyError:
                pass
        max_length = min(max_length, 30)
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width
