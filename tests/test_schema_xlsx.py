import json
from copy import deepcopy
from functools import partial
from pathlib import Path
from traceback import print_tb

import pytest

from tests.conftest import TOKEN, SCHEMAS_URL, match_uploaded_json

from elisctl.schema import download_command, xlsx
from elisctl.schema.upload import upload_command
from openpyxl import load_workbook, Workbook
from tests.conftest import API_URL
from typing import List, Optional

Matrix = List[List[Optional[str]]]

USERNAME = "test_user@rossum.ai"
PASSWORD = "secret"

schema_id = "1"
with open(Path(__file__).parent / "data" / "schema.json") as f:
    schema_content = json.load(f)

with open(Path(__file__).parent / "data" / "xlsx_dump.json") as f:
    xlsx_content = json.load(f)


@pytest.mark.runner_setup(
    env={"ELIS_URL": API_URL, "ELIS_USERNAME": USERNAME, "ELIS_PASSWORD": PASSWORD}
)
@pytest.mark.usefixtures("mock_login_request", "mock_get_schema")
class TestXlsx:
    output_file = Path("test.xlsx")

    def test_xlsx_written_correctly(self, isolated_cli_runner):
        result = isolated_cli_runner.invoke(
            download_command, [schema_id, "--format", "xlsx", "-O", str(self.output_file)]
        )
        assert not result.exit_code, print_tb(result.exc_info[2])
        workbook = load_workbook(filename=self.output_file)
        schema_sheet = workbook[xlsx.SCHEMA_SHEET_NAME]
        assert xlsx_content["schema"] == self._sheet_to_matrix(schema_sheet)
        options_sheet = workbook[xlsx.OPTIONS_SHEET_PREFIX + "test_enum"]
        assert xlsx_content["options"] == self._sheet_to_matrix(options_sheet)

    @staticmethod
    def _sheet_to_matrix(sheet) -> Matrix:
        return [list(row) for row in sheet.values]

    def test_xlsx_loaded_correctly(self, isolated_cli_runner, requests_mock):
        xlsx_schema = deepcopy(xlsx_content["schema"])
        xlsx_options = deepcopy(xlsx_content["options"])
        # rename label
        xlsx_schema[2][4] = "Invoice no"
        # add superfluous space -- should be removed
        xlsx_schema[2][5] += " "
        # delete date_issue
        del xlsx_schema[3]
        # append one option
        xlsx_options.append(["x123", "xxx"])
        self._create_xlsx(xlsx_schema, xlsx_options, self.output_file)

        expected_schema = deepcopy(schema_content)
        del expected_schema[0]["icon"]
        del expected_schema[1]["icon"]
        del expected_schema[1]["children"][0]["default_value"]
        del expected_schema[1]["children"][2]["default_value"]
        del expected_schema[1]["children"][2]["min_occurrences"]
        # rename label
        expected_schema[0]["children"][0]["label"] = "Invoice no"
        # delete date_issue
        del expected_schema[0]["children"][1]
        # append one option
        expected_schema[1]["children"][1]["options"].append({"value": "x123", "label": "xxx"})

        schema_url = f"{SCHEMAS_URL}/{schema_id}"
        requests_mock.patch(
            schema_url,
            additional_matcher=partial(match_uploaded_json, {"content": expected_schema}),
            request_headers={"Authorization": f"Token {TOKEN}"},
            status_code=200,
            json={"content": []},
        )

        result = isolated_cli_runner.invoke(
            upload_command, [schema_id, "--rewrite", "--format", "xlsx", str(self.output_file)]
        )
        assert not result.exit_code, print_tb(result.exc_info[2])

    def _create_xlsx(self, schema_content: Matrix, options_content: Matrix, filename: str):
        workbook = Workbook()
        schema_sheet = workbook.active
        schema_sheet.title = xlsx.SCHEMA_SHEET_NAME
        self._matrix_to_sheet(schema_content, schema_sheet)
        options_sheet = workbook.create_sheet(xlsx.OPTIONS_SHEET_PREFIX + "test_enum")
        self._matrix_to_sheet(options_content, options_sheet)
        workbook.save(filename)

    @staticmethod
    def _matrix_to_sheet(rows: Matrix, sheet) -> None:
        for row in rows:
            sheet.append(row)
